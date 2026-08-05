"""
parse-sap-excel.py (v2)

Reads SAP export (1.XLSX + 2.XLSX) and generates sap_lines.json
for import-sap-direct.js.

Changes from v1:
  - Extracts poleStart and poleEnd from pole names ("Opopa N N")
  - Main sections (001-499) merged into one main line entry
  - Branches (sections 500+) become separate LINES entries
  - Does not generate sap_new_voltages.json -- all VOLTAGES already in DB

Usage:
    python scripts/parse-sap-excel.py --lines 2.XLSX --poles 1.XLSX
"""

import sys
import os
import re
import json
import argparse
from collections import defaultdict, Counter

sys.stdout.reconfigure(line_buffering=True)

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

ap = argparse.ArgumentParser()
ap.add_argument("--lines",   default="2.XLSX")
ap.add_argument("--poles",   default="1.XLSX")
ap.add_argument("--out-dir", default=os.path.dirname(os.path.abspath(__file__)))
args = ap.parse_args()

LINES_FILE = args.lines
POLES_FILE = args.poles
OUT_DIR    = args.out_dir

for f in (LINES_FILE, POLES_FILE):
    if not os.path.exists(f):
        print(f"File not found: {f}")
        sys.exit(1)

BE_TO_FILIAL = {"5000": 1, "5200": 2, "5300": 3, "5400": 4}

VOLT_ID_MAP = {
    (1, "VL035"):  6, (1, "VL110"):  7, (1, "VL330"):  9,
    (2, "VL035"):  1, (2, "VL110"):  2, (2, "VL220"):  3,
    (2, "VL330"):  4, (2, "VL750"):  5,
    (3, "VL035"): 10, (3, "VL110"): 11, (3, "VL330"): 12,
    (4, "VL035"): 13, (4, "VL110"): 14, (4, "VL220"): 15, (4, "VL330"): 16,
}

RE_POLE_NUM = re.compile(r"#\s*(\d+)")

# Step 1: read lines from 2.XLSX
print(f"Reading lines: {LINES_FILE}")
wb2 = openpyxl.load_workbook(LINES_FILE, data_only=True, read_only=True)
lines_info = {}
for row in wb2.active.iter_rows(min_row=2, values_only=True):
    code   = str(row[1]).strip() if row[1] else ""
    name   = str(row[2]).strip() if row[2] else ""
    be_raw = str(row[7]).strip() if row[7] else ""
    vc_raw = str(row[11]).strip() if row[11] else ""
    if not code or be_raw not in BE_TO_FILIAL: continue
    m = re.match(r"(VL\d+)", vc_raw)
    if not m: continue
    vc = m.group(1)
    fid = BE_TO_FILIAL[be_raw]
    vid = VOLT_ID_MAP.get((fid, vc))
    if vid is None: continue
    lines_info[code] = {"name": name or code, "filialId": fid, "voltClass": vc, "voltageId": vid}
del wb2
print(f"  Lines loaded: {len(lines_info)}")

# Step 2: read poles from 1.XLSX, group by section
print(f"Reading poles: {POLES_FILE}")
wb1 = openpyxl.load_workbook(POLES_FILE, data_only=True, read_only=True)
sections = {}
total_poles = 0
for row in wb1.active.iter_rows(min_row=2, values_only=True):
    pole_name  = str(row[2]).strip() if row[2] else ""
    parent_sec = str(row[3]).strip() if row[3] else ""
    sec_name   = str(row[4]).strip() if row[4] else ""
    if not parent_sec: continue
    parts = parent_sec.split("-")
    if len(parts) < 3: continue
    line_code = f"{parts[0]}-{parts[1]}"
    if line_code not in lines_info: continue
    try: sec_num = int(parts[2])
    except ValueError: continue
    if parent_sec not in sections:
        sections[parent_sec] = {"line_code": line_code, "sec_num": sec_num, "sec_name": sec_name, "poles": []}
    norm = pole_name.replace("№", "#")
    m2 = RE_POLE_NUM.search(norm)
    if m2:
        sections[parent_sec]["poles"].append(int(m2.group(1)))
        total_poles += 1
del wb1
print(f"  Sections: {len(sections)}, poles with numbers: {total_poles}")

# Step 3: group sections by line
line_to_sections = defaultdict(list)
for sec_code, sec_data in sections.items():
    line_to_sections[sec_data["line_code"]].append((sec_code, sec_data))
print(f"  Lines with sections: {len(line_to_sections)}")

# Step 4: build output records
output_lines = []
stats = {"main": 0, "branch": 0, "no_poles": 0}

for line_code in sorted(line_to_sections.keys()):
    sec_list  = line_to_sections[line_code]
    info      = lines_info[line_code]
    main_name = info["name"]
    filial_id = info["filialId"]
    volt_id   = info["voltageId"]

    main_secs = sorted(
        [(c, d) for c, d in sec_list if d["sec_num"] < 500],
        key=lambda x: x[1]["sec_num"]
    )
    branches = sorted(
        [(c, d) for c, d in sec_list if d["sec_num"] >= 500],
        key=lambda x: x[1]["sec_num"]
    )

    if main_secs:
        all_poles = []
        for _, d in main_secs:
            all_poles.extend(d["poles"])
        if all_poles:
            ps, pe, pc = min(all_poles), max(all_poles), max(all_poles) - min(all_poles) + 1
        else:
            ps = pe = pc = None
            stats["no_poles"] += 1
        output_lines.append({"sapCode": line_code, "name": main_name, "filialId": filial_id,
            "voltageId": volt_id, "poleStart": ps, "poleEnd": pe, "poleCount": pc})
        stats["main"] += 1

    for sec_code, d in branches:
        sec_label   = d["sec_name"] if d["sec_name"] else ("Branch-" + str(d["sec_num"]))
        branch_name = main_name + " / " + sec_label
        poles = d["poles"]
        if poles:
            ps2, pe2, pc2 = min(poles), max(poles), max(poles) - min(poles) + 1
        else:
            ps2 = pe2 = pc2 = None
            stats["no_poles"] += 1
        output_lines.append({"sapCode": sec_code, "name": branch_name, "filialId": filial_id,
            "voltageId": volt_id, "poleStart": ps2, "poleEnd": pe2, "poleCount": pc2})
        stats["branch"] += 1

lines_with_secs = set(line_to_sections.keys())
for line_code in sorted(lines_info.keys()):
    if line_code in lines_with_secs: continue
    info = lines_info[line_code]
    output_lines.append({"sapCode": line_code, "name": info["name"], "filialId": info["filialId"],
        "voltageId": info["voltageId"], "poleStart": None, "poleEnd": None, "poleCount": None})
    stats["no_poles"] += 1

# Stats
FILIAL_NAMES = {1: "Gomel", 2: "Zhlobin", 3: "Mozyr", 4: "Rechitsa"}
print("\nResult:")
print(f"  Main lines: {stats['main']}")
print(f"  Branches:   {stats['branch']}")
print(f"  No poles:   {stats['no_poles']}")
print(f"  Total:      {len(output_lines)}")
by_f = Counter(l['filialId'] for l in output_lines)
for fid in sorted(by_f):
    print(f"  Filial {fid} ({FILIAL_NAMES[fid]}): {by_f[fid]}")

lines_out = os.path.join(OUT_DIR, "sap_lines.json")
with open(lines_out, "w", encoding="utf-8") as f:
    json.dump(output_lines, f, ensure_ascii=False, indent=2)
print(f"\nSaved: {lines_out}")
print("\nNext:")
print("  node scripts/import-sap-direct.js --dry-run")
print("  node scripts/import-sap-direct.js")